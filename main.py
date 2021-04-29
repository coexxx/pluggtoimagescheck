import requests
import io
import pandas as pd
import itertools
import openpyxl

from PIL import Image
from PIL import ImageChops
from openpyxl.styles import Font

def main():
  #abre a planilha plugg.to
  df = pd.read_excel('planilha.xlsx')

  resolucao = []
  dupes = []
  n = 0

  print('Corrigindo', (len(df['sku (*)'])), 'SKUs:')

  for y in range(len(df['sku (*)'])):
    print(y + 1, '-', 'SKU:', df['sku (*)'][y])

    links = [df['link_image_1'][n], df['link_image_2'][n], df['link_image_3'][n], df['link_image_4'][n], df['link_image_5'][n], df['link_image_6'][n]]

    for a, b in itertools.combinations(links, 2):
      if isinstance(a, float) == False and isinstance(b, float) == False:
        try:
          #download imgs
          link1 = requests.get(a)
          link2 = requests.get(b)

          #carrega as imagens na memória
          hash1 = io.BytesIO(link1.content)
          hash2 = io.BytesIO(link2.content)

          #abre as imagens no pillow
          img1 = Image.open(hash1)
          img2 = Image.open(hash2)

          #check de duplicidade - diferença entre os canais de cores
          diff = ImageChops.difference(img1, img2)

        except Exception as e:
          print(e)
          pass

        #se nulo é duplicada
        if diff.getbbox() == None and b not in dupes:
          dupes.append(b)

        #check de resolução da imagem
        if img1.size != (1000, 1000) and img1.size != (900, 900):
          #converte imagem em 1000x1000 via api da vtex (CASO IMG ESTEJA HOSPEDADA NA VTEX)
          resolucao_mil = a[:53] + "-1000-1000" + a[53:]

          if resolucao_mil not in resolucao and a not in dupes:
            #adiciona alteração como dicionário
            resolucao.append({a: resolucao_mil})

    n += 1

  #loop de alteração dos links com dimensões inválidas
  for produto in resolucao:
    df.replace(to_replace = produto, inplace=True)

  #exporta pra o excel
  df.to_excel('duplicadas.xlsx', index=False)
  print('planilha salva!')
  #altera a fonte para vermelho para ser excluída ao subir
  wb = openpyxl.load_workbook("duplicadas.xlsx")
  ws = wb.active

  #caminha pela planilha, encontra o valor e altera a cor da fonte para vermelho
  #ao subir algum atributo para a plugg.to com a fonte vermelha, a informação é excluída
  i = 0
  for produto in dupes:
    for r in range(1,ws.max_row+1):
        for c in range(1,ws.max_column+1):
            s = ws.cell(r,c).value
            cell = ws.cell(r,c)
            if s != None and s == produto:
                cell.font = Font(color='00FF0000')
                print(cell)
                i += 1

  wb.save('duplicadas.xlsx')

  print('Acabou!')

if __name__ == '__main__':
  main()
